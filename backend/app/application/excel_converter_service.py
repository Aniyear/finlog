"""Excel Converter Service — in-memory grouping engine.

Processes Excel files by grouping rows on a key column:
- Numeric columns → SUM
- Text columns → unique values joined by comma
No database interaction — purely stateless.
"""

from __future__ import annotations

import io
import logging
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB
MAX_PREVIEW_ROWS = 500


def _is_numeric(value: Any) -> bool:
    """Check if a value is numeric."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.replace(",", ".").replace(" ", ""))
            return True
        except (ValueError, AttributeError):
            return False
    return False


def _to_number(value: Any) -> float:
    """Convert value to float."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", ".").replace(" ", ""))
        except (ValueError, AttributeError):
            return 0.0
    return 0.0


class ExcelConverterService:
    """In-memory Excel grouping engine."""

    @staticmethod
    def preview_input(file_bytes: bytes, sheet_name: str | None = None) -> dict:
        """
        Parse Excel and return preview data.

        Returns:
            {
                "sheets": ["Sheet1", "Sheet2"],
                "current_sheet": "Sheet1",
                "columns": ["Col1", "Col2", ...],
                "sample_rows": [[val1, val2, ...], ...],
                "row_count": 150,
                "column_types": {"Col1": "text", "Col2": "numeric", ...}
            }
        """
        if len(file_bytes) > MAX_FILE_SIZE:
            raise ValueError(f"File too large. Maximum size: {MAX_FILE_SIZE // (1024*1024)} MB")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheets = wb.sheetnames
        except Exception as exc:
            raise ValueError(f"Cannot read Excel file: {exc}")

        if sheet_name and sheet_name in sheets:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title if ws else ""

        if ws is None:
            wb.close()
            raise ValueError("No active sheet found")

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            raise ValueError("File must have at least a header row and one data row")

        # First row = headers
        headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(rows[0])]

        # Data rows
        data_rows = rows[1:]
        row_count = len(data_rows)

        # Auto-detect column types by sampling first 100 rows
        sample_size = min(100, row_count)
        column_types: dict[str, str] = {}

        for col_idx, header in enumerate(headers):
            numeric_count = 0
            non_empty_count = 0
            for row_idx in range(sample_size):
                if col_idx < len(data_rows[row_idx]):
                    val = data_rows[row_idx][col_idx]
                    if val is not None and str(val).strip():
                        non_empty_count += 1
                        if _is_numeric(val):
                            numeric_count += 1

            # If >60% of non-empty values are numeric, classify as numeric
            if non_empty_count > 0 and (numeric_count / non_empty_count) > 0.6:
                column_types[header] = "numeric"
            else:
                column_types[header] = "text"

        # Sample rows for preview
        sample_rows = []
        for row in data_rows[:MAX_PREVIEW_ROWS]:
            sample_rows.append([
                str(cell) if cell is not None else ""
                for cell in row
            ])

        return {
            "sheets": sheets,
            "current_sheet": sheet_name,
            "columns": headers,
            "sample_rows": sample_rows,
            "row_count": row_count,
            "column_types": column_types,
        }

    @staticmethod
    def process(
        file_bytes: bytes,
        group_by_column: str,
        column_rules: dict[str, str],
        sheet_name: str | None = None,
    ) -> dict:
        """
        Group rows by key column and aggregate.

        Args:
            file_bytes: Raw Excel file bytes
            group_by_column: Column name to group by
            column_rules: {"col_name": "sum" | "unique_join" | "first" | "skip"}

        Returns:
            {
                "columns": [...],
                "grouped_rows": [[...], ...],
                "original_count": 150,
                "grouped_count": 42,
                "preview_rows": [[...], ...],  # first N rows
            }
        """
        if len(file_bytes) > MAX_FILE_SIZE:
            raise ValueError(f"File too large. Maximum size: {MAX_FILE_SIZE // (1024*1024)} MB")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot read Excel file: {exc}")

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if ws is None:
            wb.close()
            raise ValueError("No active sheet found")

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            raise ValueError("File must have at least a header row and one data row")

        headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(rows[0])]
        data_rows = rows[1:]

        if group_by_column not in headers:
            raise ValueError(f"Column '{group_by_column}' not found. Available: {headers}")

        group_col_idx = headers.index(group_by_column)

        # Build groups
        groups: dict[str, list[list[Any]]] = {}
        for row in data_rows:
            row_list = list(row)
            # Pad row if shorter than headers
            while len(row_list) < len(headers):
                row_list.append(None)

            key = str(row_list[group_col_idx]) if row_list[group_col_idx] is not None else ""
            if key not in groups:
                groups[key] = []
            groups[key].append(row_list)

        # Determine output columns (exclude skipped)
        output_headers = []
        output_col_indices = []
        for i, header in enumerate(headers):
            rule = column_rules.get(header, "unique_join")
            if rule != "skip":
                output_headers.append(header)
                output_col_indices.append(i)

        # Aggregate
        grouped_rows: list[list[Any]] = []
        for key, group_rows in groups.items():
            result_row: list[Any] = []
            for out_idx, col_idx in enumerate(output_col_indices):
                header = headers[col_idx]
                rule = column_rules.get(header, "unique_join")

                if header == group_by_column:
                    # Group key column — just use the key
                    result_row.append(key)
                elif rule == "sum":
                    total = sum(_to_number(row[col_idx]) for row in group_rows)
                    # Round to avoid float precision issues
                    result_row.append(round(total, 2))
                elif rule == "unique_join":
                    values = []
                    for row in group_rows:
                        val = row[col_idx]
                        val_str = str(val).strip() if val is not None else ""
                        if val_str and val_str not in values:
                            values.append(val_str)
                    result_row.append(", ".join(values))
                elif rule == "first":
                    val = group_rows[0][col_idx]
                    result_row.append(str(val) if val is not None else "")
                elif rule == "count":
                    result_row.append(len(group_rows))
                else:
                    # Default: unique_join
                    values = []
                    for row in group_rows:
                        val = row[col_idx]
                        val_str = str(val).strip() if val is not None else ""
                        if val_str and val_str not in values:
                            values.append(val_str)
                    result_row.append(", ".join(values))

            grouped_rows.append(result_row)

        return {
            "columns": output_headers,
            "grouped_rows": grouped_rows,
            "original_count": len(data_rows),
            "grouped_count": len(grouped_rows),
            "preview_rows": grouped_rows[:MAX_PREVIEW_ROWS],
        }

    @staticmethod
    def generate_output_excel(columns: list[str], rows: list[list[Any]]) -> bytes:
        """Generate a styled Excel file from grouped data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результат"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="9CA3AF"),
            right=Side(style="thin", color="9CA3AF"),
            top=Side(style="thin", color="9CA3AF"),
            bottom=Side(style="thin", color="9CA3AF"),
        )
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write headers
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        for row_data in rows:
            ws.append(row_data)

        # Style data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = data_alignment

        # Auto-width columns
        for col_idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                        if cell_len > max_length:
                            max_length = cell_len
                    except Exception:
                        pass
            adjusted_width = min(max_length + 4, 50)
            if adjusted_width < 10:
                adjusted_width = 10
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()
